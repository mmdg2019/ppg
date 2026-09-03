# -*- coding: utf-8 -*-

from openpyxl import load_workbook

from odoo import fields, models, _
from odoo.exceptions import UserError
from odoo.tools.misc import file_path


class PurchaseReturnCreateWizard(models.TransientModel):
    _name = "script.migration.purchase.return.create.wizard"
    _description = "Purchase Return Create Wizard"

    _SOURCE_FILENAME = "Purchase Return Create (stock.location).xlsx"
    _ACCOUNT_HEADERS = (
        "Loss Account",
        "Loss Acount",
        "Stock Valuation Account",
        "Stock Valuation Account (Outgoing)",
    )
    _USAGE_BY_LABEL = {
        "customer": "customer",
        "internal": "internal",
        "inventory": "inventory",
        "inventory loss": "inventory",
        "production": "production",
        "supplier": "supplier",
        "transit": "transit",
        "vendor": "supplier",
        "view": "view",
        "virtual": "view",
    }

    missing_account_summary = fields.Text(readonly=True)
    create_summary = fields.Text(readonly=True)
    skip_summary = fields.Text(readonly=True)

    def action_check(self):
        self.ensure_one()
        missing_accounts = []

        for row in self._iter_excel_rows():
            company = self._find_company(row["company"])
            account = company and self._find_account(row["account"], company)
            if self._row_requires_account(row) and not account:
                missing_accounts.append(self._row_label(row))

        self.missing_account_summary = self._format_summary("Missing Account", missing_accounts)
        self.create_summary = False
        self.skip_summary = False
        return self._reopen()

    def action_run(self):
        self.ensure_one()
        created = []
        skipped = []

        for row in self._iter_excel_rows():
            company = self._find_company(row["company"])
            usage = self._get_usage(row["location_type"])
            parent = company and self._find_parent_location(row["parent"], company)
            account = company and self._find_account(row["account"], company)

            if (
                not company
                or not row["location"]
                or not usage
                or not parent
                or (self._row_requires_account(row) and not account)
                or self._find_existing_location(row, company, parent)
            ):
                skipped.append(self._row_label_with_account(row))
                continue

            values = {
                "name": row["location"],
                "location_id": parent.id,
                "usage": usage,
                "company_id": company.id,
            }
            if account:
                values["valuation_account_id"] = account.id

            self._location_model(company).create(values)
            created.append(self._row_label_with_account(row, row["account"] if account else ""))

        self.missing_account_summary = False
        self.create_summary = self._format_summary("Create", created)
        self.skip_summary = self._format_summary("Skip", skipped)
        return self._reopen()

    def _iter_excel_rows(self):
        path = file_path(
            "script_migration_stock_config/static/data/%s" % self._SOURCE_FILENAME,
            filter_ext=(".xlsx",),
        )
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            rows = sheet.iter_rows(values_only=True)
            headers = next(rows, ())
            index_by_header = {
                self._cell_text(header): index
                for index, header in enumerate(headers)
                if self._cell_text(header)
            }
            missing_headers = [
                header
                for header in ("Location Name", "Parent Location", "Location Type", "Company")
                if header not in index_by_header
            ]
            account_header = next(
                (header for header in self._ACCOUNT_HEADERS if header in index_by_header),
                "",
            )
            if not account_header:
                missing_headers.append("Loss Account / Stock Valuation Account")
            if missing_headers:
                raise UserError(_("Missing Excel column(s): %s", ", ".join(missing_headers)))

            for values in rows:
                if not any(self._cell_text(value) for value in values):
                    continue
                yield {
                    "location": self._cell_text(
                        self._cell_at(values, index_by_header["Location Name"])
                    ),
                    "parent": self._cell_text(
                        self._cell_at(values, index_by_header["Parent Location"])
                    ),
                    "location_type": self._cell_text(
                        self._cell_at(values, index_by_header["Location Type"])
                    ),
                    "company": self._cell_text(
                        self._cell_at(values, index_by_header["Company"])
                    ),
                    "account": self._cell_text(
                        self._cell_at(values, index_by_header[account_header])
                    ),
                }
        finally:
            workbook.close()

    def _find_company(self, company_name):
        if not company_name:
            return self.env["res.company"]
        companies = self.env["res.company"].search([("name", "=", company_name)], limit=2)
        return companies if len(companies) == 1 else self.env["res.company"]

    def _find_parent_location(self, parent_name, company):
        if not parent_name:
            return self.env["stock.location"]

        Location = self._location_model(company).with_context(active_test=False)
        company_domain = ["|", ("company_id", "=", False), ("company_id", "=", company.id)]
        for field_name in ("complete_name", "name"):
            locations = Location.search([(field_name, "=", parent_name)] + company_domain)
            location = self._select_company_location(locations, company)
            if location:
                return location
        return self.env["stock.location"]

    def _select_company_location(self, locations, company):
        if len(locations) == 1:
            return locations

        company_locations = locations.filtered(lambda location: location.company_id == company)
        if len(company_locations) == 1:
            return company_locations

        shared_locations = locations.filtered(lambda location: not location.company_id)
        if len(shared_locations) == 1 and not company_locations:
            return shared_locations

        return self.env["stock.location"]

    def _find_existing_location(self, row, company, parent):
        if not row["location"]:
            return self.env["stock.location"]

        return self._location_model(company).with_context(active_test=False).search([
            ("name", "=", row["location"]),
            ("company_id", "=", company.id),
            ("location_id", "=", parent.id),
        ], limit=1)

    def _find_account(self, account_name, company):
        if not account_name:
            return self.env["account.account"]

        code, name = self._split_account(account_name)
        Account = self.env["account.account"].with_company(company).with_context(
            allowed_company_ids=company.ids,
        )
        company_domain = Account._check_company_domain(company)

        if code and name:
            accounts = Account.search(company_domain + [("code", "=", code)], limit=2)
            accounts = accounts.filtered(
                lambda account: self._same_text(account.name, name)
                or self._same_text("%s %s" % (account.code, account.name), account_name)
            )
        elif code:
            accounts = Account.search(company_domain + [("code", "=", code)], limit=2)
        else:
            accounts = Account.search(company_domain + [("name", "ilike", account_name)])
            accounts = accounts.filtered(
                lambda account: self._same_text(account.name, account_name)
            )

        return accounts if len(accounts) == 1 else self.env["account.account"]

    def _location_model(self, company):
        return self.env["stock.location"].with_company(company).with_context(
            allowed_company_ids=company.ids,
        )

    def _row_requires_account(self, row):
        return bool(row["account"]) or self._get_usage(row["location_type"]) in ("inventory", "production")

    def _get_usage(self, location_type):
        return self._USAGE_BY_LABEL.get(self._normalize_key(location_type))

    @staticmethod
    def _split_account(account_name):
        parts = account_name.split(" ", 1)
        if len(parts) == 2 and any(char.isdigit() for char in parts[0]):
            return parts[0], parts[1].strip()
        return "", account_name

    @staticmethod
    def _cell_text(value):
        if value is None:
            return ""
        if isinstance(value, float) and value.is_integer():
            return str(int(value)).strip()
        return str(value).strip()

    @classmethod
    def _same_text(cls, left, right):
        return " ".join(cls._cell_text(left).split()) == " ".join(cls._cell_text(right).split())

    @classmethod
    def _normalize_key(cls, value):
        return " ".join(cls._cell_text(value).lower().split())

    @staticmethod
    def _cell_at(values, index):
        return values[index] if index < len(values) else None

    @staticmethod
    def _row_label(row):
        return "%s - %s" % (row["company"], row["location"])

    @staticmethod
    def _row_label_with_account(row, account_name=None):
        if account_name is None:
            account_name = row["account"]
        return "%s - %s - %s" % (row["company"], row["location"], account_name)

    @staticmethod
    def _format_summary(title, records):
        lines = ["%s (Total - %s)" % (title, len(records))]
        lines.extend(records)
        return "\n".join(lines)

    def _reopen(self):
        self.ensure_one()
        view = self.env.ref(
            "script_migration_stock_config.view_script_migration_purchase_return_create_wizard_form",
            raise_if_not_found=False,
        )
        action = {
            "type": "ir.actions.act_window",
            "name": _("Purchase Return Create"),
            "res_model": self._name,
            "res_id": self.id,
            "view_mode": "form",
            "target": "new",
        }
        if view:
            action.update({"view_id": view.id, "views": [(view.id, "form")]})
        return action
