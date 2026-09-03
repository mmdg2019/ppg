# -*- coding: utf-8 -*-

from openpyxl import load_workbook

from odoo import models, _
from odoo.exceptions import UserError
from odoo.tools.misc import file_path

from .purchase_return_create_wizard import PurchaseReturnCreateWizard


class RecycleReceiptCreateWizard(models.TransientModel):
    _name = "script.migration.recycle.receipt.create.wizard"
    _inherit = "script.migration.purchase.return.create.wizard"
    _description = "Recycle Receipt Create Wizard"

    _SOURCE_FILENAME = "Recycle Receipt (stock.location).xlsx"
    _ACCOUNT_HEADERS = PurchaseReturnCreateWizard._ACCOUNT_HEADERS
    _USAGE_BY_LABEL = PurchaseReturnCreateWizard._USAGE_BY_LABEL

    def action_run(self):
        self.ensure_one()
        created = []
        skipped = []

        for row in self._iter_excel_rows():
            company = self._find_company(row["company"])
            usage = self._get_usage(row["location_type"])
            parent = company and self._find_parent_location(row["parent"], company)
            account = company and self._find_account(row["account"], company)

            if (
                not company
                or not row["location"]
                or not usage
                or (row["parent"] and not parent)
                or (self._row_requires_account(row) and not account)
                or self._find_existing_location(row, company, parent)
            ):
                skipped.append(self._row_label_with_account(row))
                continue

            values = {
                "name": row["location"],
                "usage": usage,
                "company_id": company.id,
            }
            if parent:
                values["location_id"] = parent.id
            if account:
                values["valuation_account_id"] = account.id

            self._location_model(company).create(values)
            created.append(self._row_label_with_account(row, row["account"] if account else ""))

        self.missing_account_summary = False
        self.create_summary = self._format_summary("Create", created)
        self.skip_summary = self._format_summary("Skip", skipped)
        return self._reopen()

    def _iter_excel_rows(self):
        path = file_path(
            "script_migration_stock_config/static/data/%s" % self._SOURCE_FILENAME,
            filter_ext=(".xlsx",),
        )
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            rows = sheet.iter_rows(values_only=True)
            headers = next(rows, ())
            index_by_header = {
                self._cell_text(header): index
                for index, header in enumerate(headers)
                if self._cell_text(header)
            }
            missing_headers = [
                header
                for header in ("Location Name", "Location Type", "Company")
                if header not in index_by_header
            ]
            account_header = next(
                (header for header in self._ACCOUNT_HEADERS if header in index_by_header),
                "",
            )
            if not account_header:
                missing_headers.append("Loss Account / Stock Valuation Account")
            if missing_headers:
                raise UserError(_("Missing Excel column(s): %s", ", ".join(missing_headers)))

            parent_index = index_by_header.get("Parent Location")
            for values in rows:
                if not any(self._cell_text(value) for value in values):
                    continue
                yield {
                    "location": self._cell_text(
                        self._cell_at(values, index_by_header["Location Name"])
                    ),
                    "parent": self._cell_text(self._cell_at(values, parent_index))
                    if parent_index is not None
                    else "",
                    "location_type": self._cell_text(
                        self._cell_at(values, index_by_header["Location Type"])
                    ),
                    "company": self._cell_text(
                        self._cell_at(values, index_by_header["Company"])
                    ),
                    "account": self._cell_text(
                        self._cell_at(values, index_by_header[account_header])
                    ),
                }
        finally:
            workbook.close()

    def _find_existing_location(self, row, company, parent):
        if not row["location"]:
            return self.env["stock.location"]

        return self._location_model(company).with_context(active_test=False).search([
            ("name", "=", row["location"]),
            ("company_id", "=", company.id),
            ("location_id", "=", parent.id if parent else False),
        ], limit=1)

    def _reopen(self):
        self.ensure_one()
        view = self.env.ref(
            "script_migration_stock_config.view_script_migration_recycle_receipt_create_wizard_form",
            raise_if_not_found=False,
        )
        action = {
            "type": "ir.actions.act_window",
            "name": _("Recycle Receipt"),
            "res_model": self._name,
            "res_id": self.id,
            "view_mode": "form",
            "target": "new",
        }
        if view:
            action.update({"view_id": view.id, "views": [(view.id, "form")]})
        return action
