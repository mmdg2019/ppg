# -*- coding: utf-8 -*-

from openpyxl import load_workbook

from odoo import fields, models, _
from odoo.exceptions import UserError


class ScriptMigrationOperationTypeWizard(models.TransientModel):
    _name = "script.migration.operation.type.wizard"
    _description = "Script Migration Operation Type Wizard"

    _FLOW_CONFIG = {
        "purchase_return": {
            "title": "Purchase Return",
            "excel_file": "/Users/waiyan/Downloads/Purchase Return (stock.picking.type).xlsx",
            "location_side": "destination",
            "operation_type_header": ("Operation Type",),
            "company_header": ("Company",),
            "target_location_header": ("Destination Location",),
            "expected_location": {
                "name": "Purchase Return",
                "parent_name": "Virtual Locations",
            },
            "view_xml_id": (
                "script_migration_stock_config."
                "view_script_migration_operation_type_wizard_form"
            ),
        },
    }
    _LOCATION_FIELD_BY_SIDE = {
        "source": "default_location_src_id",
        "destination": "default_location_dest_id",
    }

    config_key = fields.Selection(
        selection="_selection_config_key",
        default=lambda self: self.env.context.get("default_config_key", "purchase_return"),
        required=True,
        readonly=True,
    )
    summary = fields.Text(readonly=True)

    def _selection_config_key(self):
        return [
            (config_key, config["title"])
            for config_key, config in self._FLOW_CONFIG.items()
        ]

    def action_check(self):
        self.ensure_one()
        config = self._get_flow_config()
        location_field = self._get_location_field(config)
        updates = []
        missing_locations = []

        for row in self._iter_excel_rows(config):
            company = self._find_company(row["company"])
            if not company:
                continue

            picking_type = self._find_picking_type(row["operation_type"], company)
            target_location = self._find_expected_location(config, company)
            if not target_location:
                self._append_unique(missing_locations, self._company_label(row))
                continue

            if not picking_type:
                continue

            current_location = picking_type[location_field]
            if current_location != target_location:
                updates.append(
                    "%s - %s - %s"
                    % (
                        self._company_label(row),
                        self._operation_type_label(row),
                        self._location_label(current_location),
                    )
                )

        self.summary = self._format_summary([
            ("Update", updates),
            ("Location Missing", missing_locations),
        ])
        return self._reopen()

    def action_run(self):
        self.ensure_one()
        config = self._get_flow_config()
        location_field = self._get_location_field(config)
        updated = []
        skipped = []

        for row in self._iter_excel_rows(config):
            company = self._find_company(row["company"])
            picking_type = self._find_picking_type(row["operation_type"], company) if company else False
            target_location = self._find_expected_location(config, company) if company else False

            if not company or not picking_type or not target_location:
                self._append_unique(skipped, self._company_label(row))
                continue

            if picking_type[location_field] != target_location:
                picking_type.write({location_field: target_location.id})
                updated.append(
                    "%s - %s - %s"
                    % (
                        self._company_label(row),
                        self._operation_type_label(row),
                        self._location_label(target_location),
                    )
                )

        self.summary = self._format_summary([
            ("Update", updated),
            ("Skip", skipped),
        ])
        return self._reopen()

    def _iter_excel_rows(self, config):
        try:
            workbook = load_workbook(config["excel_file"], read_only=True, data_only=True)
        except FileNotFoundError as error:
            raise UserError(_("Excel file not found: %s") % config["excel_file"]) from error

        try:
            sheet = workbook.active
            rows = sheet.iter_rows(values_only=True)
            headers = next(rows, ())
            index_by_header = {
                self._cell_text(header): index
                for index, header in enumerate(headers)
                if self._cell_text(header)
            }
            indexes = self._get_column_indexes(config, index_by_header)

            for values in rows:
                if not any(self._cell_text(value) for value in values):
                    continue
                yield {
                    "operation_type": self._cell_text(
                        self._cell_at(values, indexes["operation_type"])
                    ),
                    "company": self._cell_text(self._cell_at(values, indexes["company"])),
                    "target_location": self._cell_text(
                        self._cell_at(values, indexes["target_location"])
                    ),
                }
        finally:
            workbook.close()

    def _get_column_indexes(self, config, index_by_header):
        column_configs = {
            "operation_type": config["operation_type_header"],
            "company": config["company_header"],
            "target_location": config["target_location_header"],
        }
        indexes = {}
        missing_columns = []
        for key, headers in column_configs.items():
            header = self._get_existing_header(headers, index_by_header)
            if header:
                indexes[key] = index_by_header[header]
            else:
                missing_columns.append(" / ".join(headers))

        if missing_columns:
            raise UserError(_("Missing Excel column(s): %s") % ", ".join(missing_columns))
        return indexes

    def _get_flow_config(self):
        self.ensure_one()
        return self._FLOW_CONFIG.get(self.config_key) or self._FLOW_CONFIG["purchase_return"]

    def _get_location_field(self, config):
        location_field = self._LOCATION_FIELD_BY_SIDE.get(config.get("location_side"))
        if not location_field:
            raise UserError(_("Unsupported Operation Type location side."))
        return location_field

    def _find_company(self, company_name):
        if not company_name:
            return self.env["res.company"]
        companies = self.env["res.company"].search([("name", "=", company_name)], limit=2)
        return companies if len(companies) == 1 else self.env["res.company"]

    def _find_picking_type(self, operation_type_name, company):
        if not operation_type_name or not company:
            return self.env["stock.picking.type"]
        picking_types = self._picking_type_model(company).search([
            ("name", "=", operation_type_name),
            ("company_id", "=", company.id),
        ], limit=2)
        return picking_types if len(picking_types) == 1 else self.env["stock.picking.type"]

    def _find_expected_location(self, config, company):
        expected_location = config["expected_location"]
        location_name = expected_location.get("name")
        parent_name = expected_location.get("parent_name")
        if not location_name or not parent_name or not company:
            return self.env["stock.location"]

        locations = self._location_model(company).search([
            ("name", "=", location_name),
            ("company_id", "=", company.id),
            ("location_id.name", "=", parent_name),
            ("location_id.company_id", "in", [False, company.id]),
        ], limit=2)
        return locations if len(locations) == 1 else self.env["stock.location"]

    def _picking_type_model(self, company):
        return self.env["stock.picking.type"].with_company(company).with_context(
            active_test=False,
            allowed_company_ids=company.ids,
        )

    def _location_model(self, company):
        return self.env["stock.location"].with_company(company).with_context(
            active_test=False,
            allowed_company_ids=company.ids,
        )

    def _reopen(self):
        self.ensure_one()
        config = self._get_flow_config()
        view = self.env.ref(config["view_xml_id"], raise_if_not_found=False)
        action = {
            "type": "ir.actions.act_window",
            "name": _(config["title"]),
            "res_model": self._name,
            "res_id": self.id,
            "view_mode": "form",
            "target": "new",
            "context": {"default_config_key": self.config_key},
        }
        if view:
            action.update({"view_id": view.id, "views": [(view.id, "form")]})
        return action

    @staticmethod
    def _get_existing_header(headers, index_by_header):
        for header in headers:
            if header in index_by_header:
                return header
        return ""

    @staticmethod
    def _cell_text(value):
        if value is None:
            return ""
        if isinstance(value, float) and value.is_integer():
            return str(int(value)).strip()
        return str(value).strip()

    @staticmethod
    def _cell_at(values, index):
        return values[index] if index < len(values) else None

    @classmethod
    def _company_label(cls, row):
        return cls._cell_text(row["company"]) or "Blank"

    @classmethod
    def _operation_type_label(cls, row):
        return cls._cell_text(row["operation_type"]) or "Blank"

    @classmethod
    def _location_label(cls, location):
        if not location:
            return "Blank"
        return cls._cell_text(location.complete_name or location.display_name or location.name)

    @staticmethod
    def _append_unique(records, value):
        if value not in records:
            records.append(value)

    @staticmethod
    def _format_summary(sections):
        lines = []
        for title, records in sections:
            if lines:
                lines.append("")
            lines.append("%s (Total - %s)" % (title, len(records)))
            lines.extend(records)
        return "\n".join(lines)
