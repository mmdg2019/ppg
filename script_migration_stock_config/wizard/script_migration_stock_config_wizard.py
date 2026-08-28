# -*- coding: utf-8 -*-

from openpyxl import load_workbook

from odoo import fields, models, _
from odoo.exceptions import UserError
from odoo.tools.misc import file_path


class ScriptMigrationStockConfigWizard(models.TransientModel):
    _name = "script.migration.stock.config.wizard"
    _description = "Script Migration Stock Config Wizard"

    _FLOW_CONFIG = {
        "damage_receipt": {
            "filename": "Damage Receipt (stock.location).xlsx",
            "account_headers": ("Loss Account", "Loss Acount"),
            "title": "Damage Receipt Loss Account",
            "view_xml_id": "script_migration_stock_config.view_script_migration_stock_config_wizard_form",
        },
        "damage_stock_consumption": {
            "filename": "Damage Stock Consumption (stock.location).xlsx",
            "account_headers": (
                "Loss Account",
                "Loss Acount",
                "Stock Valuation Account (Outgoing)",
            ),
            "title": "Damage Stock Consumption Loss Account",
            "view_xml_id": "script_migration_stock_config.view_script_migration_damage_stock_consumption_wizard_form",
        },
        "fg_damage": {
            "filename": "FG Damage (stock.location).xlsx",
            "account_headers": ("Loss Account", "Loss Acount"),
            "title": "FG Damage Loss Account",
            "view_xml_id": "script_migration_stock_config.view_script_migration_fg_damage_wizard_form",
            "update_with_account": True,
        },
        "finished_usage": {
            "filename": "Finished Usage (stock.location).xlsx",
            "account_headers": ("Loss Account", "Loss Acount"),
            "title": "Finished Usage Loss Account",
            "view_xml_id": "script_migration_stock_config.view_script_migration_finished_usage_wizard_form",
            "update_with_account": True,
        },
        "raw_usage": {
            "filename": "Raw Usage (stock.location).xlsx",
            "account_headers": ("Loss Account", "Loss Acount"),
            "title": "Raw Usage",
            "view_xml_id": "script_migration_stock_config.view_script_migration_raw_usage_wizard_form",
            "update_with_account": True,
        },
        "recycle_incoming": {
            "filename": "Recycle Incoming (stock.location).xlsx",
            "account_headers": ("Loss Account", "Loss Acount"),
            "title": "Recycle Incoming",
            "view_xml_id": "script_migration_stock_config.view_script_migration_recycle_incoming_wizard_form",
            "update_with_account": True,
        },
        "production": {
            "filename": "Production (stock.location).xlsx",
            "account_headers": ("Loss Account", "Loss Acount"),
            "title": "Production",
            "view_xml_id": "script_migration_stock_config.view_script_migration_production_wizard_form",
            "update_with_account": True,
            "skip_blank_account": True,
        },
        "inventory_adjustment": {
            "filename": "Inventory Adj (stock.location).xlsx",
            "account_headers": ("Loss Account", "Loss Acount"),
            "title": "Inventory Adjustment",
            "view_xml_id": "script_migration_stock_config.view_script_migration_inventory_adjustment_wizard_form",
            "update_with_account": True,
            "skip_blank_account": True,
        },
    }

    config_key = fields.Selection(
        [
            ("damage_receipt", "Damage Receipt"),
            ("damage_stock_consumption", "Damage Stock Consumption"),
            ("fg_damage", "FG Damage"),
            ("finished_usage", "Finished Usage"),
            ("raw_usage", "Raw Usage"),
            ("recycle_incoming", "Recycle Incoming"),
            ("production", "Production"),
            ("inventory_adjustment", "Inventory Adjustment"),
        ],
        default=lambda self: self.env.context.get("default_config_key", "damage_receipt"),
        required=True,
        readonly=True,
    )
    summary = fields.Text(readonly=True)

    def action_check(self):
        self.ensure_one()
        config = self._get_flow_config()
        blank_locations = []
        missing_accounts = []

        for row in self._iter_excel_rows():
            company = self._find_company(row["company"])
            location = company and self._find_location(row["location"], company)

            if location and not location.valuation_account_id:
                blank_locations.append(self._row_label(row))

            should_check_account = row["account"] or not config.get("skip_blank_account")
            if should_check_account and (
                not company or not self._find_account(row["account"], company)
            ):
                missing_accounts.append(self._row_label(row))

        self.summary = self._format_summary([
            ("Loss Account Blank", blank_locations),
            ("Missing Account", missing_accounts),
        ])
        return self._reopen()

    def action_run(self):
        self.ensure_one()
        config = self._get_flow_config()
        updated = []
        skipped = []

        for row in self._iter_excel_rows():
            company = self._find_company(row["company"])
            location = company and self._find_location(row["location"], company)
            if config.get("skip_blank_account") and not row["account"]:
                skipped.append(self._row_label(row))
                continue
            account = company and self._find_account(row["account"], company)

            if company and location and account:
                location.valuation_account_id = account
                if config.get("update_with_account"):
                    updated.append("%s - %s" % (self._row_label(row), self._account_label(account)))
                else:
                    updated.append(self._row_label(row))
            else:
                skipped.append(self._row_label(row))

        self.summary = self._format_summary([
            ("Update", updated),
            ("Skip", skipped),
        ])
        return self._reopen()

    def _iter_excel_rows(self):
        config = self._get_flow_config()
        path = file_path(
            "script_migration_stock_config/static/data/%s" % config["filename"],
            filter_ext=(".xlsx",),
        )
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        headers = next(rows, ())
        index_by_header = {self._cell_text(header): index for index, header in enumerate(headers)}
        missing_headers = [
            header for header in ("Location Name", "Company") if header not in index_by_header
        ]
        account_headers = config["account_headers"]
        if isinstance(account_headers, str):
            account_headers = (account_headers,)
        account_header = next(
            (header for header in account_headers if header in index_by_header),
            "",
        )
        if not account_header:
            missing_headers.append(" / ".join(account_headers))
        if missing_headers:
            raise UserError(_("Missing Excel column(s): %s", ", ".join(missing_headers)))

        for values in rows:
            yield {
                "location": self._cell_text(self._cell_at(values, index_by_header["Location Name"])),
                "company": self._cell_text(self._cell_at(values, index_by_header["Company"])),
                "account": self._cell_text(self._cell_at(values, index_by_header[account_header])),
            }

    def _find_company(self, company_name):
        if not company_name:
            return self.env["res.company"]
        companies = self.env["res.company"].search([("name", "=", company_name)], limit=2)
        return companies if len(companies) == 1 else self.env["res.company"]

    def _find_location(self, location_name, company):
        if not location_name:
            return self.env["stock.location"]
        Location = self.env["stock.location"].with_company(company).with_context(
            allowed_company_ids=company.ids,
        )
        locations = Location.search([
            ("name", "=", location_name),
            ("company_id", "=", company.id),
        ], limit=2)
        return locations if len(locations) == 1 else self.env["stock.location"]

    def _find_account(self, account_name, company):
        if not account_name:
            return self.env["account.account"]

        code, name = self._split_account(account_name)
        domain = [("company_ids", "in", company.ids)]
        Account = self.env["account.account"].with_company(company).with_context(
            allowed_company_ids=company.ids,
        )

        if code and name:
            accounts = Account.search(domain + [("code", "=", code)], limit=2)
            accounts = accounts.filtered(
                lambda account: self._same_text(account.name, name)
                or self._same_text("%s %s" % (account.code, account.name), account_name)
            )
        elif code:
            accounts = Account.search(domain + [("code", "=", code)], limit=2)
        else:
            accounts = Account.search(domain + [("name", "ilike", account_name)])
            accounts = accounts.filtered(
                lambda account: self._same_text(account.name, account_name)
            )

        return accounts if len(accounts) == 1 else self.env["account.account"]

    def _get_flow_config(self):
        self.ensure_one()
        return self._FLOW_CONFIG.get(self.config_key) or self._FLOW_CONFIG["damage_receipt"]

    @staticmethod
    def _split_account(account_name):
        parts = account_name.split(" ", 1)
        if len(parts) == 2 and any(char.isdigit() for char in parts[0]):
            return parts[0], parts[1].strip()
        return "", account_name

    @staticmethod
    def _cell_text(value):
        if value is None:
            return ""
        if isinstance(value, float) and value.is_integer():
            return str(int(value)).strip()
        return str(value).strip()

    @classmethod
    def _same_text(cls, left, right):
        return " ".join(cls._cell_text(left).split()) == " ".join(cls._cell_text(right).split())

    @staticmethod
    def _cell_at(values, index):
        return values[index] if index < len(values) else None

    @staticmethod
    def _account_label(account):
        return " ".join(value for value in (account.code, account.name) if value)

    @staticmethod
    def _row_label(row):
        return "%s - %s" % (row["company"], row["location"])

    @staticmethod
    def _format_summary(sections):
        lines = []
        for title, records in sections:
            if lines:
                lines.append("")
            lines.append("%s (Total - %s)" % (title, len(records)))
            lines.extend(records)
        return "\n".join(lines)

    def _reopen(self):
        self.ensure_one()
        config = self._get_flow_config()
        view = self.env.ref(config["view_xml_id"], raise_if_not_found=False)
        action = {
            "type": "ir.actions.act_window",
            "name": _(config["title"]),
            "res_model": self._name,
            "res_id": self.id,
            "view_mode": "form",
            "target": "new",
            "context": {"default_config_key": self.config_key},
        }
        if view:
            action.update({"view_id": view.id, "views": [(view.id, "form")]})
        return action
